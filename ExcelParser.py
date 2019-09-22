""" --------------------------------------  Last Modified on Tue, 26 Jul 2019  --------------------------------------"""

'''---------------------------------------  Author: Markos Gavalas             --------------------------------------'''


'''The following scipt is parsing particular informations from .xlsx files and put them into .csv files              '''
'''Please, close the Excel files that you want to parse before you run the program (otherwise it will give an error) '''
from openpyxl import load_workbook
import os
import numpy as np
try:
    from StringIO import StringIO
except ImportError:
    from io import StringIO
#################################################################### configs

'''
 Please insert here:
 1. The excel Table name
 2. The schema of the Table
 3. The sheets
 4. The excel sheet rows 
    (if one cell corresponds to many indices(rows) insert the min index) 
'''

config = {
    "Zufriedenheit mit dem Arbeitsplatz": {
        "schema": ["Zahlenwert von 1-10", "Anmerkungen"],
        "sheet": "Mitarbeitergespräch",
        "inter_rows": [176, 0, 45]
    },
       "Zufriedenheit mit der Betreuung durch den Vorgesetzten": {
        "schema": ["Zahlenwert von 1-10", "Anmerkungen"],
        "sheet": "Mitarbeitergespräch",
        "inter_rows": [186, 0, 45]
    }
}

stamconf = {
    "Stammdaten Stelle": {
        "schema": "KSTFachbereichAbteilung",
        "sheet": "Stammdaten",
        "inter_rows": [5, 4, 9]
    }
}
''' Please insert the directory where the .xlsx files are in. If nothing is given
 the program will search for the files that are in the directory of the script'''
xlsx_dir = "."

'''This will be the directory where the program will put the parsed csv files
it will be in parsed_files and under the date where the execution of the program happend'''

#GT_Ordner = "\\\\m-fs2\\GT-Daten\\Fachbereich\\IT-1\\IT-11\\Analytics\\Analytics_Cloud\\gavam\\QS\\"
GT_Ordner = "."
'''The program will create 2 .csv files in the directory of the script'''
'''The delimiter will be ;'''
delimiter = ";"
############################################################################
xlsxs = [f for f in os.listdir(xlsx_dir) if f.endswith("xlsx")]


"""Function that parse the information into a list
IMPORTANT: The column "Zahlenwert von 1-10" has to be either empty or a digit 
           and the column "Anmerkungen" has to be more than just a number
           otherwise the outputs will have errors
:param wb: workbook
:param SheetName: The tab name of the sheet
:param minmaxr: the row number (min=max because we just consider each time 1 row)
:param mincol: minimum column number
:param maxcol: max column number
:param listlen: The length of the list we want to get as output 
"""
def parser(wb, SheetName, minmaxr, mincol, maxcol, listlen):
    sheet = wb[SheetName]
    list_row = ['-']*listlen
    for row in sheet.iter_rows(min_row=minmaxr, max_row=minmaxr, min_col=mincol, max_col=maxcol):
        index = 0
        for cell in row:
            if listlen == 1:
                if cell.value is not None:
                    list_row[index] = " ".join(cell.value.replace(";", "").split('\n') if isinstance(cell.value,str) else cell.value)
                    break
            else:
                if cell.value is not None:
                    if str(cell.value).isdigit():
                        list_row[index] = (cell.value.replace(";", "") if isinstance(cell.value, str) else cell.value)
                    else:
                        index = 1
                        list_row[index] = " ".join(cell.value.replace(";", "").split('\n') if isinstance(cell.value, str) else cell.value)
                        break
    return list_row


#create a list of wb so that we do not read the excels more that once
wbs = [load_workbook(xlsx) for xlsx in xlsxs]



for excel_type in config.keys():
    lofl = []#list of lists
    for wb in wbs:
        lofl.append(parser(wb, config[excel_type]["sheet"],
                           config[excel_type]["inter_rows"][0],
                           config[excel_type]["inter_rows"][1],
                           config[excel_type]["inter_rows"][2], 2)
                    +
                    parser(wb, stamconf["Stammdaten Stelle"]["sheet"],
                           stamconf["Stammdaten Stelle"]["inter_rows"][0],
                           stamconf["Stammdaten Stelle"]["inter_rows"][1],
                           stamconf["Stammdaten Stelle"]["inter_rows"][2], 1)
                    )
    csv_name = excel_type.replace(" ", "_") + '.csv'
    header = delimiter.join(config[excel_type]["schema"] + [stamconf["Stammdaten Stelle"]["schema"]])
    np.savetxt(GT_Ordner + csv_name, np.array(lofl), delimiter=delimiter, header=header, comments='', fmt='%s')




